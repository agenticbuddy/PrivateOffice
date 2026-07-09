"""New-document template tests — the xlsx first-sheet name is localized per creator locale."""
from io import BytesIO

from openpyxl import load_workbook

from app.services import templates


def _first_sheet_name(data: bytes) -> str:
    return load_workbook(BytesIO(data)).sheetnames[0]


def test_sheet_title_localized():
    assert templates._sheet_title("ru") == "Лист1"
    assert templates._sheet_title("en") == "Sheet1"
    assert templates._sheet_title("de") == "Tabelle1"
    assert templates._sheet_title("pt-BR") == "Planilha1"
    assert templates._sheet_title("pt") == "Folha1"
    assert templates._sheet_title("ru-RU") == "Лист1"  # region falls back to language
    assert templates._sheet_title("xx") == "Sheet1"    # unknown → English
    assert templates._sheet_title(None) == "Sheet1"


def test_blank_xlsx_first_sheet_follows_locale():
    assert _first_sheet_name(templates.blank_document("xlsx", "ru")) == "Лист1"
    assert _first_sheet_name(templates.blank_document("xlsx", "en")) == "Sheet1"
    # no locale → neutral English default
    assert _first_sheet_name(templates.blank_document("xlsx")) == "Sheet1"


def test_docx_pptx_ignore_locale():
    # smoke: non-xlsx builders still produce valid bytes and ignore the locale arg
    assert templates.blank_document("docx", "ru")[:2] == b"PK"
    assert templates.blank_document("pptx", "ru")[:2] == b"PK"
